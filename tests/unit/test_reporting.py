from __future__ import annotations

import csv
import json
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from cse_financial_etl.reporting.dashboard import generate_run_dashboard
from cse_financial_etl.reporting.excel import generate_excel

AS_OF = date(2026, 9, 5)
PERIOD = date(2025, 12, 31)
RUN_ID = "test-run-dashboard"


def _seed_outputs(root: Path) -> Path:
    outputs = root / "outputs"
    api = root / "data" / "raw" / "api"
    outputs.mkdir(parents=True)
    api.mkdir(parents=True)
    facts_path = outputs / f"normalized_facts_{AS_OF.isoformat()}.csv"
    with facts_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "issuer_name",
                "symbol",
                "period_end",
                "metric_code",
                "metric_type",
                "normalized_value",
                "status",
                "entity_scope",
                "source_page",
                "extraction_method",
                "certainty_band",
                "source_line",
                "duration_months",
                "validation_status",
                "review_status",
            ],
        )
        writer.writeheader()
        writer.writerow(
            {
                "issuer_name": "Acme PLC",
                "symbol": "ACM.N0000",
                "period_end": PERIOD.isoformat(),
                "metric_code": "PAT",
                "metric_type": "MONETARY_ABSOLUTE",
                "normalized_value": "1000",
                "status": "EXTRACTED",
                "entity_scope": "COMPANY",
                "source_page": "2",
                "extraction_method": "LAYOUT_TEXT",
                "certainty_band": "HIGH",
                "source_line": "Profit for the period 1,000",
                "duration_months": "3",
                "validation_status": "PASSED",
                "review_status": "APPROVED",
            }
        )
        writer.writerow(
            {
                "issuer_name": "Acme PLC",
                "symbol": "ACM.N0000",
                "period_end": PERIOD.isoformat(),
                "metric_code": "OPERATING_PROFIT",
                "metric_type": "MONETARY_ABSOLUTE",
                "normalized_value": "",
                "status": "SOURCE_CONFIRMED_NOT_REPORTED",
                "entity_scope": "COMPANY",
                "source_page": "",
                "extraction_method": "",
                "certainty_band": "NONE",
                "source_line": "",
                "duration_months": "",
                "validation_status": "",
                "review_status": "",
            }
        )
        writer.writerow(
            {
                "issuer_name": "Acme PLC",
                "symbol": "ACM.N0000",
                "period_end": PERIOD.isoformat(),
                "metric_code": "TOTAL_LIABILITIES",
                "metric_type": "MONETARY_ABSOLUTE",
                "normalized_value": "",
                "status": "NOT_FOUND_BY_PARSER",
                "entity_scope": "COMPANY",
                "source_page": "",
                "extraction_method": "",
                "certainty_band": "NONE",
                "source_line": "",
                "duration_months": "",
                "validation_status": "",
                "review_status": "",
            }
        )
        writer.writerow(
            {
                "issuer_name": "Acme PLC",
                "symbol": "ACM.N0000",
                "period_end": PERIOD.isoformat(),
                "metric_code": "PBT",
                "metric_type": "MONETARY_ABSOLUTE",
                "normalized_value": "999",
                "status": "EXTRACTED",
                "entity_scope": "COMPANY",
                "source_page": "2",
                "extraction_method": "LAYOUT_TEXT",
                "certainty_band": "HIGH",
                "source_line": "Profit before tax 999",
                "duration_months": "3",
                "validation_status": "FAILED",
                "review_status": "REVIEW",
            }
        )
    prices_path = outputs / f"quarter_end_prices_{AS_OF.isoformat()}.csv"
    with prices_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "issuer_name",
                "symbol",
                "period_end",
                "value",
                "source_method",
                "status",
                "source_page",
            ],
        )
        writer.writeheader()
        writer.writerow(
            {
                "issuer_name": "Acme PLC",
                "symbol": "ACM.N0000",
                "period_end": PERIOD.isoformat(),
                "value": "12.5",
                "source_method": "FILING_LAYOUT",
                "status": "EXTRACTED",
                "source_page": "5",
            }
        )
        writer.writerow(
            {
                "issuer_name": "Blank PLC",
                "symbol": "BLN.N0000",
                "period_end": PERIOD.isoformat(),
                "value": "",
                "source_method": "",
                "status": "HISTORICAL_PRICE_NOT_AVAILABLE",
                "source_page": "",
            }
        )
    reviews_path = outputs / f"review_queue_{AS_OF.isoformat()}.csv"
    with reviews_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["issuer_name", "symbol", "period_end", "metric_code", "reason", "detail"],
        )
        writer.writeheader()
        writer.writerow(
            {
                "issuer_name": "Acme PLC",
                "symbol": "ACM.N0000",
                "period_end": PERIOD.isoformat(),
                "metric_code": "NAVPS",
                "reason": "NOT_FOUND_BY_PARSER",
                "detail": "No NAVPS line",
            }
        )
    (api / f"market_cap_{AS_OF.isoformat()}.json").write_text(
        json.dumps(
            [
                {
                    "company_name": "Acme PLC",
                    "symbol": "ACM.N0000",
                    "price": 12.5,
                    "issued_quantity": 1000,
                    "market_capitalization": 12500,
                    "market_cap_percentage": 100.0,
                }
            ]
        ),
        encoding="utf-8",
    )
    (outputs / f"golden_validation_{AS_OF.isoformat()}.json").write_text(
        json.dumps(
            {
                "sample_size": 1,
                "passed": 1,
                "failed": 0,
                "accuracy": 1.0,
                "by_metric": {"PAT": {"sample_size": 1, "passed": 1, "failed": 0, "accuracy": 1.0}},
            }
        ),
        encoding="utf-8",
    )
    return outputs


def test_workbook_contains_only_snapshot_sheet(tmp_path: Path) -> None:
    _seed_outputs(tmp_path)
    path = generate_excel(tmp_path, AS_OF, [PERIOD], RUN_ID)
    workbook = load_workbook(path)
    assert workbook.sheetnames == [f"Snapshot_{AS_OF.isoformat()}"]
    assert RUN_ID in str(workbook.active["A2"].value)
    assert "dashboard.html" in str(workbook.active["G2"].value)
    # Header rows stay frozen; identity columns must scroll freely.
    assert workbook.active.freeze_panes == "A5"
    # Failed validation must not display the extraction status as the cell value.
    pbt_header = None
    for cell in workbook.active[4]:
        if cell.value == "PBT":
            pbt_header = cell.column
            break
    assert pbt_header is not None
    assert workbook.active.cell(row=5, column=pbt_header).value == "VALIDATION_FAILED"


def test_run_dashboard_is_self_contained(tmp_path: Path) -> None:
    _seed_outputs(tmp_path)
    run_dir = tmp_path / "outputs" / "runs" / f"{AS_OF.isoformat()}_{RUN_ID}"
    path = generate_run_dashboard(
        tmp_path,
        AS_OF,
        [PERIOD],
        RUN_ID,
        run_dir,
        {"run_status": "COMPLETED_WITH_REVIEW", "pipeline_error_count": 0, "issuer_count": 1},
    )
    html = path.read_text(encoding="utf-8")
    assert path.name == "dashboard.html"
    assert RUN_ID in html
    assert "Review queue" in html
    assert "NOT_FOUND_BY_PARSER" in html
    assert "Honest misses" in html
    assert "Parser gaps" in html
    assert "SOURCE_CONFIRMED_NOT_REPORTED" in html
    assert "HISTORICAL_PRICE_NOT_AVAILABLE" in html
    assert "LOW_CERTAINTY only" in html
    assert "https://" not in html
    latest = tmp_path / "outputs" / f"run_dashboard_{AS_OF.isoformat()}.html"
    assert latest.exists()
