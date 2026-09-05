from __future__ import annotations

import csv
import json
from collections import Counter
from collections.abc import Iterable
from datetime import date
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from cse_financial_etl.sources.historical_prices import resolve_quarter_end_price

VISIBLE_METRICS: tuple[tuple[str, str], ...] = (
    ("PAT", "PAT"),
    ("PBT", "PBT"),
    ("EPS_SELECTED", "EPS (Diluted Preferred)"),
    ("NAVPS", "NAVPS"),
    ("OPERATING_PROFIT", "Operating Profit"),
    ("TOTAL_EQUITY", "Total Equity"),
    ("TOTAL_ASSETS", "Total Assets"),
    ("TOTAL_LIABILITIES", "Total Liabilities"),
    ("TOP_LINE", "Revenue / Gross Income"),
)
DERIVED_HEADERS = (
    "Market Price at Quarter End",
    "Debt to Equity (x)",
    "ROE (Quarter)",
    "ROA (Quarter)",
    "NPM (Quarter)",
)
ACCEPTED_STATUSES = {"EXTRACTED", "EXTRACTED_DERIVED"}
BALANCE_SHEET_CODES = {
    "TOTAL_ASSETS",
    "TOTAL_EQUITY",
    "TOTAL_LIABILITIES",
    "DEBT_TO_EQUITY",
    "ROE",
    "ROA",
}
BALANCE_SHEET_REASONS = {
    "BALANCE_SHEET_RECONCILIATION_FAILED",
    "BALANCE_SHEET_SANITY_FAILED",
}


def _read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _number(value: str | None) -> float | None:
    try:
        return float(value) if value not in (None, "") else None
    except (TypeError, ValueError):
        return None


def _market_snapshot_date(path: Path) -> date | None:
    if path.parent.name.startswith("date="):
        try:
            return date.fromisoformat(path.parent.name.removeprefix("date="))
        except ValueError:
            return None
    stem = path.stem
    if not stem.startswith("market_cap_"):
        return None
    try:
        return date.fromisoformat(stem.removeprefix("market_cap_")[:10])
    except ValueError:
        return None


def _rows_from_snapshot(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        return payload if isinstance(payload, list) else []
    if path.suffix.lower() == ".parquet":
        return pl.read_parquet(path).to_dicts()
    return []


def _ranks_from_rows(rows: list[dict[str, Any]]) -> dict[str, int]:
    ranked = sorted(
        rows,
        key=lambda row: (
            row.get("market_capitalization") is not None,
            row.get("market_capitalization") or -1,
        ),
        reverse=True,
    )
    return {
        row["symbol"]: index
        for index, row in enumerate(
            (item for item in ranked if (item.get("market_capitalization") or 0) > 0),
            start=1,
        )
        if row.get("symbol")
    }


def _previous_ranks(project_root: Path, as_of_date: date) -> dict[str, int]:
    """Use the latest prior market snapshot, including archived same-date runs."""

    live = project_root / "data" / "raw" / "api" / f"market_cap_{as_of_date.isoformat()}.json"
    candidates: list[tuple[date, float, Path]] = []

    def consider(path: Path, *, allow_same_date: bool) -> None:
        snapshot_date = _market_snapshot_date(path)
        if snapshot_date is None:
            return
        if snapshot_date > as_of_date:
            return
        if snapshot_date == as_of_date and not allow_same_date:
            return
        try:
            mtime = path.stat().st_mtime
        except OSError:
            return
        candidates.append((snapshot_date, mtime, path))

    api_dir = project_root / "data" / "raw" / "api"
    if api_dir.exists():
        for path in api_dir.glob("market_cap_*.json"):
            if path.name == live.name:
                continue
            consider(path, allow_same_date=False)
    history_dir = api_dir / "history"
    if history_dir.exists():
        for path in history_dir.glob("market_cap_*.json"):
            consider(path, allow_same_date=True)
    silver_dir = project_root / "data" / "silver" / "market_snapshots"
    if silver_dir.exists():
        for date_dir in silver_dir.glob("date=*"):
            parts = [path for path in date_dir.glob("*.parquet") if path.is_file()]
            if not parts:
                continue
            latest = max(parts, key=lambda item: item.stat().st_mtime)
            consider(latest, allow_same_date=True)
    if not candidates:
        return {}
    _snapshot_date, _mtime, path = max(candidates, key=lambda item: (item[0], item[1]))
    return _ranks_from_rows(_rows_from_snapshot(path))


def generate_excel(
    project_root: Path, as_of_date: date, periods: Iterable[date], run_id: str
) -> Path:
    periods = tuple(periods)
    facts = _read_csv(project_root / "outputs" / f"normalized_facts_{as_of_date.isoformat()}.csv")
    prices = _read_csv(
        project_root / "outputs" / f"quarter_end_prices_{as_of_date.isoformat()}.csv"
    )
    reviews = _read_csv(project_root / "outputs" / f"review_queue_{as_of_date.isoformat()}.csv")
    market: list[dict[str, Any]] = json.loads(
        (
            project_root / "data" / "raw" / "api" / f"market_cap_{as_of_date.isoformat()}.json"
        ).read_text(encoding="utf-8")
    )

    fact_map = {(row["issuer_name"], row["period_end"], row["metric_code"]): row for row in facts}
    price_map = {(row["symbol"], row["period_end"]): row for row in prices}
    period_reason: dict[tuple[str, str], str] = {}
    for row in reviews:
        period_reason.setdefault(
            (row.get("issuer_name", ""), row.get("period_end", "")),
            row.get("reason") or "REVIEW_REQUIRED",
        )
    balance_sheet_failed = {
        (row.get("issuer_name", ""), row.get("period_end", ""))
        for row in reviews
        if row.get("reason") in BALANCE_SHEET_REASONS
    }
    previous_ranks = _previous_ranks(project_root, as_of_date)

    def value_or_reason(issuer: str, period: date, code: str) -> float | str:
        if (
            code in BALANCE_SHEET_CODES
            and (
                issuer,
                period.isoformat(),
            )
            in balance_sheet_failed
        ):
            return "BALANCE_SHEET_REVIEW"
        row = fact_map.get((issuer, period.isoformat(), code))
        value = _number(row.get("normalized_value")) if row else None
        if row and row.get("status") in ACCEPTED_STATUSES and value is not None:
            return value
        return (
            (row.get("status") if row else None)
            or period_reason.get((issuer, period.isoformat()))
            or "NOT_REPORTED"
        )

    def quarter_ratio(issuer: str, period: date, denominator_code: str) -> float | str:
        profit = value_or_reason(issuer, period, "PAT")
        denominator = value_or_reason(issuer, period, denominator_code)
        if not isinstance(profit, float) or not isinstance(denominator, float):
            return "INSUFFICIENT_INPUT"
        if denominator <= 0:
            return (
                "ZERO_EQUITY" if denominator_code == "TOTAL_EQUITY" else "NON_POSITIVE_DENOMINATOR"
            )
        return profit / denominator

    def derived_metric(issuer: str, period: date, code: str, fallback: float | str) -> float | str:
        value = value_or_reason(issuer, period, code)
        if value != "NOT_REPORTED":
            return value
        return fallback

    ranked = sorted(
        market,
        key=lambda row: (
            row.get("market_capitalization") is not None,
            row.get("market_capitalization") or -1,
        ),
        reverse=True,
    )
    rank_by_symbol = {
        row["symbol"]: index
        for index, row in enumerate(
            (item for item in ranked if (item.get("market_capitalization") or 0) > 0), start=1
        )
    }

    wb = Workbook()
    wb.remove(wb.active)
    readme = wb.create_sheet("README")
    snapshot = wb.create_sheet(f"Snapshot_{as_of_date.isoformat()}")
    checks = wb.create_sheet("Checks")
    accuracy = wb.create_sheet("Accuracy_Certainty")
    audit = wb.create_sheet("Audit_Lineage")
    price_audit = wb.create_sheet("Price_Lineage")
    review = wb.create_sheet("Review_Queue")
    coverage = wb.create_sheet("Extraction_Coverage")
    failure = wb.create_sheet("Failure_Analysis")
    definitions = wb.create_sheet("Metric_Definitions")

    navy, blue, mid_blue, green = "17365D", "1F4E78", "5B9BD5", "548235"
    light_blue, amber, white = "D9EAF7", "FFF2CC", "FFFFFF"
    thin = Side(style="thin", color="D9E2F3")

    def title(ws: Any, cell_range: str, text: str) -> None:
        ws.merge_cells(cell_range)
        cell = ws[cell_range.split(":")[0]]
        cell.value = text
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=white, bold=True, size=16)

    def style_header(cells: Any, fill: str = blue) -> None:
        for cell in cells:
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(color=white, bold=True, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)

    title(readme, "A1:H1", "CSE Financial Data Platform — Production ETL Output")
    readme_rows = (
        ("Workbook version", "3.1"),
        ("Run ID", run_id),
        ("Market snapshot", as_of_date.isoformat()),
        ("Displayed periods", " | ".join(period.isoformat() for period in periods)),
        ("Snapshot grain", "One row per exact listed security symbol"),
        (
            "Financial grain",
            "Standalone Company/Bank, exact three-month quarter for flows, as-at for stocks",
        ),
        (
            "Q4 rule",
            "Only an explicitly reported standalone 3M/4Q flow is publishable; cumulative/FY deltas are never used",
        ),
        (
            "EPS rule",
            "Diluted EPS when reported; otherwise basic. Both values remain in Audit_Lineage",
        ),
        (
            "Quarter ratios",
            "ROE (Quarter) = PAT / equity, ROA (Quarter) = PAT / assets, NPM (Quarter) = PAT / revenue or gross income; same quarter only; no TTM",
        ),
        (
            "Quarter price",
            "Filing first, then official CSE history, then last trade on or before quarter end. Never use a later live snapshot as a historical price",
        ),
        ("Missing values", "Typed reason text; no value is guessed or silently converted to zero"),
        ("Source", "Official Colombo Stock Exchange market data and quarterly PDF filings"),
    )
    for row_index, (label, value) in enumerate(readme_rows, start=3):
        readme.cell(row_index, 1, label)
        readme.cell(row_index, 2, value)
        readme.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=8)
        readme.cell(row_index, 1).fill = PatternFill("solid", fgColor=blue)
        readme.cell(row_index, 1).font = Font(color=white, bold=True)
        readme.cell(row_index, 2).alignment = Alignment(wrap_text=True)
    readme.column_dimensions["A"].width = 24
    for column in "BCDEFGH":
        readme.column_dimensions[column].width = 18

    columns_per_period = 14
    total_columns = 10 + len(periods) * columns_per_period
    title(
        snapshot,
        f"A1:{get_column_letter(total_columns)}1",
        "CSE Market Capitalization and Standalone Financial Snapshot",
    )
    snapshot.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    snapshot.cell(2, 1, f"Market as of: {as_of_date.isoformat()}")
    snapshot.merge_cells(start_row=2, start_column=7, end_row=2, end_column=total_columns)
    snapshot.cell(
        2, 7, "Absolute monetary values normalized to LKR; per-share amounts remain LKR/share"
    )
    for cell in snapshot[2]:
        cell.fill = PatternFill("solid", fgColor=light_blue)
        cell.font = Font(color=navy, bold=True)

    snapshot.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)
    snapshot.cell(3, 1, "Ranking and Market Data")
    for column in range(1, 11):
        snapshot.cell(3, column).fill = PatternFill("solid", fgColor=navy)
        snapshot.cell(3, column).font = Font(color=white, bold=True)
    fills = (blue, mid_blue, green)
    for index, period in enumerate(periods):
        start = 11 + index * columns_per_period
        snapshot.merge_cells(
            start_row=3, start_column=start, end_row=3, end_column=start + columns_per_period - 1
        )
        snapshot.cell(3, start, f"Financial period ended {period.isoformat()}")
        for column in range(start, start + columns_per_period):
            snapshot.cell(3, column).fill = PatternFill("solid", fgColor=fills[index % len(fills)])
            snapshot.cell(3, column).font = Font(color=white, bold=True)

    base_headers = (
        "Previous Rank",
        "Current Rank",
        "Rank Change",
        "Company Name",
        "Symbol",
        "Current Price (LKR)",
        "Issued Quantity",
        "Market Capitalization (LKR)",
        "Market Cap Share",
        "Market Data Status",
    )
    headers = list(base_headers)
    for _ in periods:
        headers.extend(header for _, header in VISIBLE_METRICS)
        headers.extend(DERIVED_HEADERS)
    for column, header in enumerate(headers, start=1):
        snapshot.cell(4, column, header)
    style_header(snapshot[4])

    for row_index, security in enumerate(ranked, start=5):
        market_cap = security.get("market_capitalization")
        current_rank = rank_by_symbol.get(security["symbol"], "UNRANKED")
        previous_rank = previous_ranks.get(security["symbol"])
        if previous_rank is None:
            previous_display: int | str = "NEW"
            rank_change: int | str = "NEW"
        elif isinstance(current_rank, int):
            previous_display = previous_rank
            rank_change = previous_rank - current_rank
        else:
            previous_display = previous_rank
            rank_change = "NEW"
        base_values = (
            previous_display,
            current_rank,
            rank_change,
            security["company_name"],
            security["symbol"],
            security.get("price") if security.get("price") is not None else "NO_PRICE",
            security.get("issued_quantity")
            if security.get("issued_quantity") is not None
            else "NOT_REPORTED",
            market_cap if market_cap is not None else "NO_PRICE",
            (security.get("market_cap_percentage") or 0) / 100 if market_cap else "NO_PRICE",
            "OK" if market_cap and market_cap > 0 else "NO_PRICE",
        )
        for column, value in enumerate(base_values, start=1):
            snapshot.cell(row_index, column, value)

        issuer, symbol = security["company_name"], security["symbol"]
        for period_index, period in enumerate(periods):
            start = 11 + period_index * columns_per_period
            values: dict[str, float | str] = {}
            for offset, (code, _header) in enumerate(VISIBLE_METRICS):
                values[code] = value_or_reason(issuer, period, code)
                snapshot.cell(row_index, start + offset, values[code])
            price_row = price_map.get((symbol, period.isoformat()))
            price = (
                _number(price_row.get("value"))
                if price_row and price_row.get("status") == "EXTRACTED"
                else None
            )
            if price is None:
                resolved = resolve_quarter_end_price(project_root, symbol, period)
                if resolved is not None:
                    price, _price_date, _method = resolved
            snapshot.cell(
                row_index,
                start + 9,
                price
                if price is not None
                else (price_row.get("status") if price_row else "HISTORICAL_PRICE_NOT_AVAILABLE"),
            )
            liabilities, equity = values["TOTAL_LIABILITIES"], values["TOTAL_EQUITY"]
            if isinstance(liabilities, float) and isinstance(equity, float):
                debt_fallback: float | str = liabilities / equity if equity else "ZERO_EQUITY"
            else:
                debt_fallback = "INSUFFICIENT_INPUT"
            snapshot.cell(
                row_index,
                start + 10,
                derived_metric(issuer, period, "DEBT_TO_EQUITY", debt_fallback),
            )
            snapshot.cell(
                row_index,
                start + 11,
                derived_metric(
                    issuer, period, "ROE", quarter_ratio(issuer, period, "TOTAL_EQUITY")
                ),
            )
            snapshot.cell(
                row_index,
                start + 12,
                derived_metric(
                    issuer, period, "ROA", quarter_ratio(issuer, period, "TOTAL_ASSETS")
                ),
            )
            snapshot.cell(
                row_index,
                start + 13,
                derived_metric(issuer, period, "NPM", quarter_ratio(issuer, period, "TOP_LINE")),
            )

        for cell in snapshot[row_index]:
            cell.border = Border(bottom=Side(style="hair", color="D9E2F3"))
            cell.font = Font(size=9)

    last_row = 4 + len(ranked)
    snapshot.freeze_panes = "F5"
    snapshot.auto_filter.ref = f"A4:{get_column_letter(total_columns)}{last_row}"
    snapshot.sheet_view.showGridLines = False
    widths = {1: 14, 2: 13, 3: 13, 4: 35, 5: 15, 6: 14, 7: 17, 8: 23, 9: 15, 10: 18}
    for column, width in widths.items():
        snapshot.column_dimensions[get_column_letter(column)].width = width
    for column in range(11, total_columns + 1):
        snapshot.column_dimensions[get_column_letter(column)].width = 18
    snapshot.row_dimensions[4].height = 52
    for row in range(5, last_row + 1):
        snapshot.cell(row, 6).number_format = "#,##0.00;[Red](#,##0.00);-"
        snapshot.cell(row, 7).number_format = "#,##0;[Red](#,##0);-"
        snapshot.cell(row, 8).number_format = "#,##0;[Red](#,##0);-"
        snapshot.cell(row, 9).number_format = "0.00%"
        for period_index in range(len(periods)):
            start = 11 + period_index * columns_per_period
            for offset in (0, 1, 4, 5, 6, 7, 8):
                snapshot.cell(row, start + offset).number_format = "#,##0;[Red](#,##0);-"
            for offset in (2, 3, 9, 10):
                snapshot.cell(row, start + offset).number_format = "0.00;[Red](0.00);-"
            for offset in (11, 12, 13):
                snapshot.cell(row, start + offset).number_format = "0.00%"
    data_range = f"K5:{get_column_letter(total_columns)}{last_row}"
    snapshot.conditional_formatting.add(
        data_range, FormulaRule(formula=["ISTEXT(K5)"], fill=PatternFill("solid", fgColor=amber))
    )

    title(checks, "A1:F1", "Pipeline Quality Checks")
    checks.append([])
    checks.append(("Check", "Actual", "Expected", "Difference", "Status", "Notes"))
    style_header(checks[3])
    unique_issuers = len({row["company_name"] for row in market})
    display_facts = [row for row in facts if row["period_end"] in {p.isoformat() for p in periods}]
    extracted = sum(row.get("status") in ACCEPTED_STATUSES for row in display_facts)
    price_extracted = sum(row.get("status") == "EXTRACTED" for row in prices)
    fact_slots = len(display_facts)
    price_slots = len(prices)
    for check in (
        (
            "Listed securities",
            len(market),
            "Live CSE universe",
            "",
            "OK" if market else "FAIL",
            "Security symbol is the market key.",
        ),
        (
            "Unique issuers",
            unique_issuers,
            "Live CSE universe",
            "",
            "OK" if unique_issuers else "FAIL",
            "Financial facts are reused across share classes.",
        ),
        (
            "Display-period facts extracted",
            extracted,
            fact_slots,
            extracted - fact_slots,
            "OK" if extracted == fact_slots else "REVIEW",
            "Includes both EPS variants and EPS_SELECTED.",
        ),
        (
            "Quarter-end prices extracted",
            price_extracted,
            price_slots,
            price_extracted - price_slots,
            "OK" if price_extracted == price_slots else "REVIEW",
            "Only source-backed filing, CSE historical, or last-trade-on-or-before-quarter-end prices are counted.",
        ),
        (
            "Review items",
            len(reviews),
            0,
            len(reviews),
            "OK" if not reviews else "REVIEW",
            "Typed exceptions remain visible.",
        ),
        ("Pipeline status", "Completed", "Completed", 0, "OK", f"Run {run_id}"),
    ):
        checks.append(check)
    for column, width in zip("ABCDEF", (32, 18, 20, 14, 14, 62), strict=True):
        checks.column_dimensions[column].width = width

    title(accuracy, "A1:K1", "Extraction Accuracy, Certainty and Coverage")
    accuracy.append([])
    accuracy.append(
        (
            "Metric",
            "Total",
            "Extracted",
            "Coverage",
            "High",
            "Medium",
            "Low",
            "No score",
            "Mean certainty",
            "Measured accuracy",
            "Validated sample",
        )
    )
    style_header(accuracy[3])
    validation_path = project_root / "outputs" / f"golden_validation_{as_of_date.isoformat()}.json"
    validation = (
        json.loads(validation_path.read_text(encoding="utf-8")) if validation_path.exists() else {}
    )
    validation_by_metric = validation.get("by_metric", {})
    fact_metrics = [code for code, _label in VISIBLE_METRICS if code != "EPS_SELECTED"]
    for metric in ["ALL", *fact_metrics]:
        subset = (
            facts if metric == "ALL" else [row for row in facts if row["metric_code"] == metric]
        )
        extracted_rows = [row for row in subset if row.get("status") in ACCEPTED_STATUSES]
        bands = Counter(row.get("certainty_band") or "NONE" for row in subset)
        scores = [
            score for row in subset if (score := _number(row.get("overall_certainty"))) is not None
        ]
        validation_row = validation if metric == "ALL" else validation_by_metric.get(metric, {})
        accuracy.append(
            (
                metric,
                len(subset),
                len(extracted_rows),
                len(extracted_rows) / len(subset) if subset else None,
                bands["HIGH"],
                bands["MEDIUM"],
                bands["LOW"],
                bands["NONE"],
                sum(scores) / len(scores) if scores else None,
                validation_row.get("accuracy"),
                validation_row.get("sample_size", 0),
            )
        )
    start_reason = accuracy.max_row + 3
    accuracy.cell(start_reason, 1, "Open review reasons")
    accuracy.cell(start_reason, 1).font = Font(bold=True, color=navy, size=12)
    accuracy.cell(start_reason + 1, 1, "Reason")
    accuracy.cell(start_reason + 1, 2, "Count")
    style_header(accuracy[start_reason + 1][:2])
    for reason, count in Counter(row.get("reason") or "UNKNOWN" for row in reviews).most_common():
        accuracy.append((reason, count))
    for row_index in range(4, 4 + len(fact_metrics) + 1):
        accuracy.cell(row_index, 4).number_format = "0.0%"
        accuracy.cell(row_index, 9).number_format = "0.0%"
        accuracy.cell(row_index, 10).number_format = "0.0%"
    for column, width in zip(
        "ABCDEFGHIJK", (24, 12, 12, 14, 10, 10, 10, 12, 16, 19, 18), strict=True
    ):
        accuracy.column_dimensions[column].width = width
    accuracy.freeze_panes = "A4"

    def append_dict_sheet(ws: Any, rows: list[dict[str, str]]) -> None:
        headers = list(rows[0].keys()) if rows else ["Status", "Detail"]
        ws.append(headers)
        style_header(ws[1])
        if rows:
            for row in rows:
                ws.append([row.get(header, "") for header in headers])
        else:
            ws.append(["OK", "No rows"])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        for index, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(index)].width = (
                55 if header in {"source_line", "source_url", "filing_title", "local_path"} else 20
            )

    append_dict_sheet(audit, facts)
    append_dict_sheet(price_audit, prices)
    append_dict_sheet(review, reviews)

    title(coverage, "A1:D1", "Extraction Coverage")
    coverage.append([])
    coverage.append(("Metric", "Total", "Extracted", "Coverage"))
    style_header(coverage[3])
    for row_index in range(4, accuracy.max_row + 1):
        metric = accuracy.cell(row_index, 1).value
        if metric in {None, "Open review reasons", "Reason"}:
            break
        coverage.append(
            (
                metric,
                accuracy.cell(row_index, 2).value,
                accuracy.cell(row_index, 3).value,
                accuracy.cell(row_index, 4).value,
            )
        )
    for column, width in zip("ABCD", (24, 12, 12, 14), strict=True):
        coverage.column_dimensions[column].width = width
    for row in coverage.iter_rows(min_row=4, min_col=4, max_col=4):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.0%"

    title(failure, "A1:B1", "Failure Analysis")
    failure.append([])
    failure.append(("Reason", "Count"))
    style_header(failure[3])
    status_counts = Counter(row.get("status") or "UNKNOWN" for row in facts)
    reason_counts = Counter(row.get("reason") or "UNKNOWN" for row in reviews)
    failure.append(("Fact statuses", ""))
    for status, count in status_counts.most_common():
        failure.append((status, count))
    failure.append(("Review reasons", ""))
    for reason, count in reason_counts.most_common():
        failure.append((reason, count))
    failure.column_dimensions["A"].width = 42
    failure.column_dimensions["B"].width = 12

    definition_rows = (
        ("Metric", "Type", "Definition", "Scaling"),
        (
            "PAT / PBT / Operating Profit / Top Line",
            "FLOW",
            "Standalone Company/Bank exact three-month quarter; audited compatible cumulative delta when necessary",
            "Detected currency and scale",
        ),
        (
            "EPS Basic",
            "PER_SHARE",
            "Exact three-month basic EPS retained internally",
            "Never inherit statement scale",
        ),
        (
            "EPS Diluted",
            "PER_SHARE",
            "Exact three-month diluted EPS retained internally",
            "Never inherit statement scale",
        ),
        (
            "EPS Selected",
            "PER_SHARE",
            "Diluted when reported, otherwise basic",
            "Never inherit statement scale",
        ),
        (
            "Assets / Equity / Liabilities",
            "STOCK",
            "Standalone Company/Bank balance at period end",
            "Detected currency and scale",
        ),
        (
            "Total Liabilities",
            "STOCK",
            "Explicit total, or derived as Assets minus Equity with lineage",
            "Normalized LKR",
        ),
        (
            "Quarter-end Price",
            "PER_SHARE",
            "Exact security class; filing first, official history fallback",
            "LKR/share",
        ),
        (
            "Debt to Equity",
            "RATIO",
            "Total Liabilities / Total Equity",
            "Multiple (x), not percentage",
        ),
        ("ROE (Quarter)", "RATIO", "Same-quarter PAT / Total Equity", "Percentage"),
        ("ROA (Quarter)", "RATIO", "Same-quarter PAT / Total Assets", "Percentage"),
        (
            "NPM (Quarter)",
            "RATIO",
            "Same-quarter PAT / revenue or gross income",
            "Percentage",
        ),
        (
            "Q4",
            "RULE",
            "Explicit standalone three-month/4Q preferred; compatible cumulative-only flows may use an audited delta; EPS is never derived",
            "Not applicable",
        ),
    )
    for row in definition_rows:
        definitions.append(row)
    style_header(definitions[1])
    for column, width in zip("ABCD", (35, 20, 78, 35), strict=True):
        definitions.column_dimensions[column].width = width

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical="center",
                        wrap_text=cell.alignment.wrap_text or ws.title != snapshot.title,
                    )

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    workbook_dir = project_root / "outputs" / "workbooks"
    workbook_dir.mkdir(parents=True, exist_ok=True)
    output_path = workbook_dir / f"CSE_Financial_Snapshot_{as_of_date.isoformat()}.xlsx"
    wb.save(output_path)
    compat_path = project_root / "outputs" / f"CSE_Financial_Snapshot_{as_of_date.isoformat()}.xlsx"
    if compat_path != output_path:
        from shutil import copy2

        copy2(output_path, compat_path)
    return output_path
